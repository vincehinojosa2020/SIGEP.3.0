# -*- coding: utf-8 -*-
"""
SIGEP - Sistema Integrado de Gestao de Exploracao e Producao
PetroNac - Petrolera Nacional S.A.
Versao 2.4.1 - Modulo Principal do Servidor API
Desenvolvido por: TechBrazil Consultoria LTDA (2016-2018)
Contrato: CT-2016/0847-PETRONAC
Ultima modificacao: 15/03/2018 por Carlos (terceirizado)
TODO: migrar para Django 3.x quando possivel (pendente desde 2020)
"""

from dotenv import load_dotenv
from pathlib import Path
import os
import sys

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

sys.path.insert(0, str(ROOT_DIR))

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import uuid
import jwt
import bcrypt
import io

# Configuracao MongoDB
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Aplicacao principal
app = FastAPI(title="SIGEP API", version="2.4.1")
api_router = APIRouter(prefix="/api")

# ============================================================
# Configuracoes de Autenticacao
# NOTA: chave JWT hardcoded como fallback - Carlos disse que ia mudar (2017)
# ============================================================
JWT_SECRET = os.environ.get('JWT_SECRET', 'chave_secreta_sigep_2018')
JWT_ALGORITHM = "HS256"


def hash_password(password):
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')


def verify_password(plain, hashed):
    return bcrypt.checkpw(plain.encode('utf-8'), hashed.encode('utf-8'))


def create_token(user_id, username, role):
    payload = {
        'sub': user_id,
        'username': username,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(hours=24)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request):
    token = None
    auth_header = request.headers.get('Authorization', '')
    if auth_header.startswith('Bearer '):
        token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail='Nao autenticado')
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.usuarios.find_one({'id': payload['sub']}, {'_id': 0})
        if not user:
            raise HTTPException(status_code=401, detail='Usuario nao encontrado')
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail='Token expirado')
    except:
        raise HTTPException(status_code=401, detail='Token invalido')


# ============================================================
# Modelos Pydantic
# ============================================================
class LoginRequest(BaseModel):
    username: str
    password: str


class FaunaCreate(BaseModel):
    especie: str
    data_observacao: str
    plataforma: str
    coordenadas_lat: float
    coordenadas_lon: float
    observador: str
    notas: str = ""


class FaunaUpdate(BaseModel):
    especie: Optional[str] = None
    data_observacao: Optional[str] = None
    plataforma: Optional[str] = None
    coordenadas_lat: Optional[float] = None
    coordenadas_lon: Optional[float] = None
    observador: Optional[str] = None
    notas: Optional[str] = None


class TelemetriaPayload(BaseModel):
    sensor_id: str
    tipo: str
    valor: float
    unidade: str
    timestamp: Optional[str] = None
    duto_id: Optional[str] = None


# ============================================================
# ROTAS DE AUTENTICACAO
# ============================================================
@api_router.post("/auth/login")
async def login(req: LoginRequest):
    user = await db.usuarios.find_one({'username': req.username}, {'_id': 0})
    if not user:
        raise HTTPException(status_code=401, detail='Credenciais invalidas')
    if not verify_password(req.password, user['password_hash']):
        raise HTTPException(status_code=401, detail='Credenciais invalidas')
    token = create_token(user['id'], user['username'], user.get('role', 'operador'))
    return {
        'token': token,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'nome': user['nome'],
            'role': user.get('role', 'operador'),
            'email': user.get('email', '')
        }
    }


@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return {
        'id': user['id'],
        'username': user['username'],
        'nome': user['nome'],
        'role': user.get('role', 'operador'),
        'email': user.get('email', ''),
        'cargo': user.get('cargo', ''),
        'departamento': user.get('departamento', '')
    }


# ============================================================
# ROTAS DE POCOS
# ============================================================
@api_router.get("/pocos")
async def listar_pocos(user=Depends(get_current_user)):
    pocos = await db.pocos.find({}, {'_id': 0}).to_list(100)
    return pocos


@api_router.get("/pocos/{poco_id}")
async def obter_poco(poco_id: str, user=Depends(get_current_user)):
    poco = await db.pocos.find_one({'id': poco_id}, {'_id': 0})
    if not poco:
        raise HTTPException(status_code=404, detail='Poco nao encontrado')
    return poco


# ============================================================
# ROTAS DE PRODUCAO
# ============================================================
@api_router.get("/producao/resumo")
async def resumo_producao(user=Depends(get_current_user)):
    pipeline = [
        {'$sort': {'data': -1}},
        {'$group': {'_id': '$poco_id', 'ultimo': {'$first': '$$ROOT'}}}
    ]
    resultados = await db.producao.aggregate(pipeline).to_list(100)

    total_barris = 0
    total_gas = 0
    soma_agua = 0
    soma_pressao = 0
    count = 0
    registros = []

    for r in resultados:
        doc = r['ultimo']
        doc.pop('_id', None)
        total_barris += doc.get('barris_por_dia', 0)
        total_gas += doc.get('gas_m3_dia', 0)
        soma_agua += doc.get('corte_agua_pct', 0)
        soma_pressao += doc.get('pressao_psi', 0)
        count += 1
        registros.append(doc)

    # Buscar nomes dos pocos
    pocos = await db.pocos.find({}, {'_id': 0, 'id': 1, 'nome': 1, 'bacia': 1, 'status': 1}).to_list(100)
    poco_map = {p['id']: p for p in pocos}
    for reg in registros:
        info = poco_map.get(reg.get('poco_id'), {})
        reg['poco_nome'] = info.get('nome', 'N/A')
        reg['bacia'] = info.get('bacia', 'N/A')
        reg['poco_status'] = info.get('status', 'N/A')

    return {
        'total_barris_dia': round(total_barris, 1),
        'total_gas_m3_dia': round(total_gas, 1),
        'media_corte_agua': round(soma_agua / max(count, 1), 1),
        'media_pressao_psi': round(soma_pressao / max(count, 1), 1),
        'total_pocos_ativos': count,
        'registros': registros
    }


@api_router.get("/producao")
async def listar_producao(
    poco_id: Optional[str] = None,
    limit: int = Query(default=50, le=500),
    user=Depends(get_current_user)
):
    filtro = {}
    if poco_id:
        filtro['poco_id'] = poco_id
    producao = await db.producao.find(filtro, {'_id': 0}).sort('data', -1).to_list(limit)
    return producao


# ============================================================
# ROTAS DE DUTOS
# ============================================================
@api_router.get("/dutos")
async def listar_dutos(user=Depends(get_current_user)):
    dutos = await db.dutos.find({}, {'_id': 0}).to_list(50)
    return dutos


@api_router.get("/dutos/{duto_id}")
async def obter_duto(duto_id: str, user=Depends(get_current_user)):
    duto = await db.dutos.find_one({'id': duto_id}, {'_id': 0})
    if not duto:
        raise HTTPException(status_code=404, detail='Duto nao encontrado')
    return duto


@api_router.get("/leituras/duto/{duto_id}")
async def leituras_duto(
    duto_id: str,
    limit: int = Query(default=50, le=500),
    user=Depends(get_current_user)
):
    leituras = await db.leituras_duto.find(
        {'duto_id': duto_id}, {'_id': 0}
    ).sort('timestamp', -1).to_list(limit)
    return leituras


# ============================================================
# TELEMETRIA SCADA
# ============================================================
@api_router.post("/telemetria/api")
async def receber_telemetria(payload: TelemetriaPayload):
    doc = payload.dict()
    doc['id'] = str(uuid.uuid4())
    if not doc.get('timestamp'):
        doc['timestamp'] = datetime.now(timezone.utc).isoformat()
    doc['processado'] = True
    doc['recebido_em'] = datetime.now(timezone.utc).isoformat()
    await db.telemetria.insert_one(doc)
    doc.pop('_id', None)
    return {'status': 'ok', 'mensagem': 'Dados recebidos com sucesso', 'id': doc['id']}


@api_router.get("/telemetria/docs")
async def telemetria_docs():
    return {
        'nome': 'API de Telemetria SCADA - SIGEP',
        'versao': '1.2',
        'descricao': 'Interface para recepcao de dados telemetricos de sensores SCADA instalados em plataformas e dutos offshore.',
        'endpoints': [
            {
                'metodo': 'POST',
                'url': '/api/telemetria/api',
                'descricao': 'Enviar leitura de sensor SCADA',
                'autenticacao': 'Nenhuma (protocolo legado)',
                'corpo': {
                    'sensor_id': 'string (obrigatorio) - Ex: SENS-P52-PRESS-001',
                    'tipo': 'string (obrigatorio) - pressao|vazao|temperatura',
                    'valor': 'number (obrigatorio)',
                    'unidade': 'string (obrigatorio) - Ex: PSI, m3/h, C',
                    'timestamp': 'string ISO 8601 (opcional)',
                    'duto_id': 'string (opcional)'
                },
                'codigos_http': {'200': 'Sucesso', '422': 'Dados invalidos'}
            }
        ],
        'exemplo_curl': "curl -X POST /api/telemetria/api -H 'Content-Type: application/json' -d '{\"sensor_id\": \"SENS-P52-PRESS-001\", \"tipo\": \"pressao\", \"valor\": 245.7, \"unidade\": \"PSI\"}'"
    }


# ============================================================
# CONFORMIDADE
# ============================================================
@api_router.get("/conformidade")
async def listar_relatorios(user=Depends(get_current_user)):
    relatorios = await db.conformidade.find({}, {'_id': 0}).sort('data_geracao', -1).to_list(50)
    return relatorios


@api_router.get("/conformidade/{relatorio_id}/pdf")
async def exportar_pdf(relatorio_id: str, user=Depends(get_current_user)):
    relatorio = await db.conformidade.find_one({'id': relatorio_id}, {'_id': 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail='Relatorio nao encontrado')

    producao = await db.producao.find(
        {'data': {'$gte': relatorio['periodo_inicio'], '$lte': relatorio['periodo_fim']}},
        {'_id': 0}
    ).to_list(5000)

    pocos = await db.pocos.find({}, {'_id': 0, 'id': 1, 'nome': 1}).to_list(100)
    poco_map = {p['id']: p['nome'] for p in pocos}

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    titulo = ParagraphStyle('TituloPetronac', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#1a2332'))

    elements = []
    elements.append(Paragraph('PETROLERA NACIONAL S.A. - PetroNac', titulo))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph('Relatorio de Conformidade ANP', styles['Heading2']))
    elements.append(Paragraph(f'Tipo: {relatorio["tipo"]}', styles['Normal']))
    elements.append(Paragraph(f'Periodo: {relatorio["periodo_inicio"]} a {relatorio["periodo_fim"]}', styles['Normal']))
    elements.append(Paragraph(f'Data de Geracao: {relatorio["data_geracao"]}', styles['Normal']))
    elements.append(Paragraph(f'Status: {relatorio["status"]}', styles['Normal']))
    elements.append(Spacer(1, 16))
    elements.append(Paragraph('Resumo de Producao por Poco', styles['Heading3']))
    elements.append(Spacer(1, 8))

    # Agrupar por poco
    por_poco = {}
    for p in producao:
        pid = p.get('poco_id', 'N/A')
        if pid not in por_poco:
            por_poco[pid] = {'barris': [], 'gas': [], 'agua': [], 'pressao': []}
        por_poco[pid]['barris'].append(p.get('barris_por_dia', 0))
        por_poco[pid]['gas'].append(p.get('gas_m3_dia', 0))
        por_poco[pid]['agua'].append(p.get('corte_agua_pct', 0))
        por_poco[pid]['pressao'].append(p.get('pressao_psi', 0))

    data_tabela = [['Poco', 'Med. bpd', 'Med. Gas m3/d', 'Med. Agua %', 'Med. PSI']]
    for pid, vals in list(por_poco.items())[:15]:
        nome = poco_map.get(pid, pid[:16])
        n_barris = len(vals['barris']) or 1
        n_gas = len(vals['gas']) or 1
        n_agua = len(vals['agua']) or 1
        n_pressao = len(vals['pressao']) or 1
        data_tabela.append([
            nome,
            f'{sum(vals["barris"])/n_barris:.0f}',
            f'{sum(vals["gas"])/n_gas:.0f}',
            f'{sum(vals["agua"])/n_agua:.1f}',
            f'{sum(vals["pressao"])/n_pressao:.0f}'
        ])

    if len(data_tabela) > 1:
        tabela = Table(data_tabela, colWidths=[4*cm, 2.5*cm, 3*cm, 2.5*cm, 2.5*cm])
        tabela.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a2332')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ]))
        elements.append(tabela)

    elements.append(Spacer(1, 24))
    elements.append(Paragraph('Documento gerado automaticamente pelo SIGEP v2.4.1', styles['Normal']))
    elements.append(Paragraph('Agencia Nacional do Petroleo, Gas Natural e Biocombustiveis - ANP', styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="relatorio_anp_{relatorio_id[:8]}.pdf"'}
    )


@api_router.get("/conformidade/{relatorio_id}/excel")
async def exportar_excel(relatorio_id: str, user=Depends(get_current_user)):
    relatorio = await db.conformidade.find_one({'id': relatorio_id}, {'_id': 0})
    if not relatorio:
        raise HTTPException(status_code=404, detail='Relatorio nao encontrado')

    producao = await db.producao.find(
        {'data': {'$gte': relatorio['periodo_inicio'], '$lte': relatorio['periodo_fim']}},
        {'_id': 0}
    ).to_list(5000)

    pocos = await db.pocos.find({}, {'_id': 0, 'id': 1, 'nome': 1}).to_list(100)
    poco_map = {p['id']: p['nome'] for p in pocos}

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatorio ANP'
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A2332', end_color='1A2332', fill_type='solid')

    ws.merge_cells('A1:F1')
    ws['A1'] = 'PETROLERA NACIONAL S.A. - PetroNac'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Relatorio de Conformidade ANP - {relatorio["tipo"]}'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A3'] = f'Periodo: {relatorio["periodo_inicio"]} a {relatorio["periodo_fim"]}'
    ws['A4'] = f'Status: {relatorio["status"]}'

    headers = ['Poco', 'Data', 'Barris/dia', 'Gas m3/dia', 'Corte Agua %', 'Pressao PSI']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, p in enumerate(producao[:500], 7):
        ws.cell(row=i, column=1, value=poco_map.get(p.get('poco_id', ''), p.get('poco_id', '')))
        ws.cell(row=i, column=2, value=p.get('data', ''))
        ws.cell(row=i, column=3, value=p.get('barris_por_dia', 0))
        ws.cell(row=i, column=4, value=p.get('gas_m3_dia', 0))
        ws.cell(row=i, column=5, value=p.get('corte_agua_pct', 0))
        ws.cell(row=i, column=6, value=p.get('pressao_psi', 0))

    for col_cells in ws.columns:
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 2, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="relatorio_anp_{relatorio_id[:8]}.xlsx"'}
    )


# ============================================================
# FAUNA
# ============================================================
@api_router.get("/fauna")
async def listar_fauna(user=Depends(get_current_user)):
    fauna = await db.fauna.find({}, {'_id': 0}).sort('data_observacao', -1).to_list(100)
    return fauna


@api_router.post("/fauna")
async def criar_observacao(obs: FaunaCreate, user=Depends(get_current_user)):
    doc = obs.dict()
    doc['id'] = str(uuid.uuid4())
    doc['criado_por'] = user['username']
    doc['criado_em'] = datetime.now(timezone.utc).isoformat()
    await db.fauna.insert_one(doc)
    doc.pop('_id', None)
    return doc


@api_router.put("/fauna/{obs_id}")
async def atualizar_observacao(obs_id: str, obs: FaunaUpdate, user=Depends(get_current_user)):
    update_data = {k: v for k, v in obs.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail='Nenhum dado para atualizar')
    result = await db.fauna.update_one({'id': obs_id}, {'$set': update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail='Observacao nao encontrada')
    updated = await db.fauna.find_one({'id': obs_id}, {'_id': 0})
    return updated


@api_router.delete("/fauna/{obs_id}")
async def deletar_observacao(obs_id: str, user=Depends(get_current_user)):
    result = await db.fauna.delete_one({'id': obs_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Observacao nao encontrada')
    return {'status': 'ok', 'mensagem': 'Observacao removida'}


# ============================================================
# USUARIOS
# ============================================================
@api_router.get("/usuarios")
async def listar_usuarios(user=Depends(get_current_user)):
    usuarios = await db.usuarios.find({}, {'_id': 0, 'password_hash': 0}).to_list(100)
    return usuarios


# ============================================================
# Configuracao de Rotas e CORS
# ============================================================
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# Seed de dados e startup
# ============================================================
from seed_data import gerar_dados_seed


@app.on_event("startup")
async def startup_event():
    count = await db.pocos.count_documents({})
    if count == 0:
        print("[SIGEP] Populando banco de dados com dados iniciais...")
        dados = gerar_dados_seed()
        for colecao, nome in [
            ('pocos', 'pocos'),
            ('producao', 'registros de producao'),
            ('dutos', 'dutos'),
            ('leituras_duto', 'leituras de duto'),
            ('conformidade', 'relatorios de conformidade'),
            ('fauna', 'observacoes de fauna')
        ]:
            if dados.get(colecao):
                await db[colecao].insert_many(dados[colecao])
                print(f"  -> {len(dados[colecao])} {nome} inseridos")
        print("[SIGEP] Seed concluido!")
    else:
        print("[SIGEP] Banco ja populado.")

    # Seed admin
    admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
    admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
    existing_admin = await db.usuarios.find_one({'username': admin_username})
    if not existing_admin:
        await db.usuarios.insert_one({
            'id': str(uuid.uuid4()),
            'username': admin_username,
            'password_hash': hash_password(admin_password),
            'nome': 'Administrador SIGEP',
            'email': 'admin@petronac.com.br',
            'role': 'admin',
            'cargo': 'Administrador do Sistema',
            'departamento': 'TI - Tecnologia da Informacao',
            'criado_em': datetime.now(timezone.utc).isoformat()
        })
        print("[SIGEP] Usuario admin criado")
    elif not verify_password(admin_password, existing_admin.get('password_hash', '')):
        await db.usuarios.update_one(
            {'username': admin_username},
            {'$set': {'password_hash': hash_password(admin_password)}}
        )

    # Seed outros usuarios
    outros = [
        {'username': 'carlos.silva', 'nome': 'Carlos Alberto da Silva', 'email': 'carlos.silva@petronac.com.br', 'role': 'engenheiro', 'cargo': 'Engenheiro de Producao', 'departamento': 'E&P'},
        {'username': 'maria.santos', 'nome': 'Maria Fernanda Santos', 'email': 'maria.santos@petronac.com.br', 'role': 'geologo', 'cargo': 'Geologa Senior', 'departamento': 'Geologia'},
        {'username': 'joao.oliveira', 'nome': 'Joao Pedro Oliveira', 'email': 'joao.oliveira@petronac.com.br', 'role': 'operador', 'cargo': 'Operador de Plataforma', 'departamento': 'Operacoes'},
        {'username': 'ana.costa', 'nome': 'Ana Beatriz Costa', 'email': 'ana.costa@petronac.com.br', 'role': 'ambiental', 'cargo': 'Analista Ambiental', 'departamento': 'Meio Ambiente'},
        {'username': 'roberto.lima', 'nome': 'Roberto Carlos Lima', 'email': 'roberto.lima@petronac.com.br', 'role': 'engenheiro', 'cargo': 'Engenheiro de Dutos', 'departamento': 'Transporte'},
    ]
    for u in outros:
        exists = await db.usuarios.find_one({'username': u['username']})
        if not exists:
            u['id'] = str(uuid.uuid4())
            u['password_hash'] = hash_password('petronac2018')
            u['criado_em'] = datetime.now(timezone.utc).isoformat()
            await db.usuarios.insert_one(u)

    # Gravar credenciais de teste
    try:
        os.makedirs('/app/memory', exist_ok=True)
        with open('/app/memory/test_credentials.md', 'w') as f:
            f.write('# Credenciais de Teste SIGEP\n\n')
            f.write('## Admin\n')
            f.write(f'- Username: {admin_username}\n')
            f.write(f'- Password: {admin_password}\n')
            f.write('- Role: admin\n\n')
            f.write('## Outros Usuarios (senha: petronac2018)\n')
            f.write('- carlos.silva\n- maria.santos\n- joao.oliveira\n- ana.costa\n- roberto.lima\n\n')
            f.write('## Endpoints\n')
            f.write('- POST /api/auth/login\n- GET /api/auth/me\n')
            f.write('- GET /api/pocos\n- GET /api/producao/resumo\n')
            f.write('- GET /api/dutos\n- GET /api/conformidade\n')
            f.write('- GET /api/fauna\n- GET /api/usuarios\n')
    except:
        pass

    print("[SIGEP] Sistema inicializado com sucesso!")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
